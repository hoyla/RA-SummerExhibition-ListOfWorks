"""Importer for Artists' Index spreadsheets.

Expected columns: Title, First Name, Last Name, Quals, Company, Address 1, Cat Nos
"""

from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy.orm import Session
from difflib import get_close_matches
from typing import Dict, List, Optional, Tuple
import re
import unicodedata
import uuid as _uuid

from backend.app.models.import_model import Import
from backend.app.models.index_artist_model import IndexArtist
from backend.app.models.index_cat_number_model import IndexCatNumber
from backend.app.models.index_override_model import IndexArtistOverride
from backend.app.models.audit_log_model import AuditLog
from backend.app.models.validation_warning_model import ValidationWarning


# ---------------------------------------------------------------------------
# RA member detection
# ---------------------------------------------------------------------------

# Tokens that indicate RA membership (case-insensitive).  Must match as
# whole words within the quals string.
RA_MEMBER_TOKENS = {
    "RA",
    "PRA",
    "PPRA",
    "HON RA",
    "HONRA",
    "RA ELECT",
    "EX OFFICIO",
}

_RA_PATTERN = re.compile(
    r"\b(?:"
    + "|".join(re.escape(t) for t in sorted(RA_MEMBER_TOKENS, key=len, reverse=True))
    + r")\b",
    re.IGNORECASE,
)


def is_ra_member(quals: Optional[str]) -> bool:
    """Return True if the quals string contains an RA-type designation."""
    if not quals:
        return False
    return bool(_RA_PATTERN.search(quals))


# ---------------------------------------------------------------------------
# Sort key
# ---------------------------------------------------------------------------


def _strip_accents(s: str) -> str:
    """Remove combining diacritical marks for sort-key purposes."""
    nfkd = unicodedata.normalize("NFKD", s)
    return "".join(c for c in nfkd if not unicodedata.combining(c))


def build_sort_key(last_name: Optional[str], first_name: Optional[str]) -> str:
    """Build a lowercase, accent-stripped sort key.

    If there is only a first name (e.g. "Assemble"), it is treated as the
    primary sort token.
    """
    primary = last_name or first_name or ""
    secondary = (first_name or "") if last_name else ""
    raw = f"{primary} {secondary}".strip().lower()
    return _strip_accents(raw)


# ---------------------------------------------------------------------------
# Cat-number parsing
# ---------------------------------------------------------------------------


def parse_cat_nos(raw: Optional[str]) -> List[int]:
    """Parse a semicolon-or-comma separated string of catalogue numbers.

    Returns a list of integers, ignoring any non-numeric tokens.
    """
    if raw is None:
        return []
    raw_str = str(raw).strip()
    if not raw_str:
        return []
    # Split on ; or ,
    tokens = re.split(r"[;,]", raw_str)
    nums: List[int] = []
    for t in tokens:
        t = t.strip()
        if t.isdigit():
            nums.append(int(t))
    return nums


# ---------------------------------------------------------------------------
# Multi-artist name parsing
# ---------------------------------------------------------------------------

# Prefix patterns indicating the last_name field contains a second artist
_SECOND_ARTIST_PREFIX = re.compile(
    r"^(?:and|&)\s+",
    re.IGNORECASE,
)


def parse_multi_artist(
    first_name: Optional[str],
    last_name: Optional[str],
    quals: Optional[str],
) -> Optional[dict]:
    """Detect and parse multi-artist entries.

    When the last_name field starts with "and " or "& ", it means the
    spreadsheet author placed the second artist there and packed the primary
    artist's full name (possibly with embedded quals) into first_name.

    Returns None if the pattern is not detected, or a dict with keys:
        first_name, last_name, quals, second_artist
    representing the parsed primary artist + second artist suffix.
    """
    if not last_name or not _SECOND_ARTIST_PREFIX.match(str(last_name).strip()):
        return None

    second_artist = str(last_name).strip()
    raw_primary = str(first_name).strip() if first_name else ""

    if not raw_primary:
        # No first name to parse — can't extract a surname
        return None

    # Strip known quals from the end of the primary name
    extracted_quals: list[str] = []
    remaining = raw_primary
    while True:
        m = _QUAL_IN_NAME_PATTERN.search(remaining)
        if not m:
            break
        # Only strip if it's at the end (after trimming)
        suffix = remaining[m.start() :].strip()
        # Check the match is at the end of the remaining string
        after_match = remaining[m.end() :].strip()
        if _QUAL_IN_NAME_PATTERN.sub("", after_match).strip() == "":
            # Everything from this match to the end is quals
            extracted_quals.append(remaining[m.start() :].strip())
            remaining = remaining[: m.start()].strip()
            break
        else:
            # Qual in the middle — just extract this one token
            extracted_quals.insert(0, m.group(0))
            remaining = (remaining[: m.start()] + remaining[m.end() :]).strip()

    # Split the remainder into first name + surname
    # Assume the last word is the surname
    words = remaining.split()
    if not words:
        return None

    parsed_last = words[-1]
    parsed_first = " ".join(words[:-1]) if len(words) > 1 else None

    # Merge extracted quals with existing quals
    all_quals_parts = []
    if extracted_quals:
        all_quals_parts.extend(extracted_quals)
    if quals and str(quals).strip():
        all_quals_parts.append(str(quals).strip())
    merged_quals = " ".join(all_quals_parts) if all_quals_parts else quals

    return {
        "first_name": parsed_first,
        "last_name": parsed_last,
        "quals": merged_quals,
        "second_artist": second_artist,
    }


# ---------------------------------------------------------------------------
# Company detection
# ---------------------------------------------------------------------------


def detect_company(
    first_name: Optional[str],
    last_name: Optional[str],
    quals: Optional[str],
) -> bool:
    """Heuristic: entry is likely a company when there is a last-name value
    but no first-name value and no qualifications."""
    has_first = bool(first_name and str(first_name).strip())
    has_last = bool(last_name and str(last_name).strip())
    has_quals = bool(quals and str(quals).strip())
    # Company if last name only, no first name
    # (If quals are present it might be a single-name individual like "Assemble RA")
    return has_last and not has_first and not has_quals


# ---------------------------------------------------------------------------
# Multi-name detection
# ---------------------------------------------------------------------------

# Words that typically separate two people's names
_MULTI_NAME_SEPARATORS = re.compile(
    r"(?:\band\b|\bwith\b|\s&\s)",
    re.IGNORECASE,
)


def detect_multi_name(first_name: Optional[str], last_name: Optional[str]) -> bool:
    """Return True if the name fields appear to contain more than one person."""
    for field in (first_name, last_name):
        if field and _MULTI_NAME_SEPARATORS.search(str(field)):
            return True
    return False


# ---------------------------------------------------------------------------
# Quals-in-name detection
# ---------------------------------------------------------------------------

# Common qualification/honorific tokens that might appear misplaced in name fields
_KNOWN_QUAL_TOKENS = {
    "RA",
    "PRA",
    "PPRA",
    "HON RA",
    "HONRA",
    "RA ELECT",
    "EX OFFICIO",
    "OBE",
    "CBE",
    "MBE",
    "DBE",
    "KBE",
    "GBE",
    "CH",
    "KCVO",
    "GCVO",
    "CVO",
    "KCB",
    "GCB",
    "DCB",
    "FRS",
    "FRSA",
    "FRIAS",
    "FRIBA",
    "RIBA",
    "RDI",
    "QPM",
    "QC",
    "KC",
}

_QUAL_IN_NAME_PATTERN = re.compile(
    r"\b(?:"
    + "|".join(re.escape(t) for t in sorted(_KNOWN_QUAL_TOKENS, key=len, reverse=True))
    + r")\b",
    re.IGNORECASE,
)


def detect_quals_in_name(
    first_name: Optional[str], last_name: Optional[str]
) -> Optional[str]:
    """If a known qualification token appears in a name field, return it.

    Returns the first matched token or None.
    """
    for field in (first_name, last_name):
        if not field:
            continue
        m = _QUAL_IN_NAME_PATTERN.search(str(field))
        if m:
            return m.group(0)
    return None


# ---------------------------------------------------------------------------
# Column headers
# ---------------------------------------------------------------------------

REQUIRED_COLUMNS = {"Last Name", "Cat Nos"}
KNOWN_COLUMNS = {
    "Title",
    "First Name",
    "Last Name",
    "Quals",
    "Company",
    "Address 1",
    "Cat Nos",
}


class IndexImportError(Exception):
    """Raised when the uploaded file cannot be imported."""

    pass


def _validate_headers(headers: List[str]) -> List[str]:
    """Validate spreadsheet headers. Returns warning strings for missing
    optional columns.  Raises IndexImportError for fatal problems."""
    found = {h for h in headers if h}
    if not found:
        raise IndexImportError(
            "The spreadsheet has no column headers in row 1. "
            "Expected columns: " + ", ".join(sorted(KNOWN_COLUMNS))
        )

    missing_required = REQUIRED_COLUMNS - found
    if missing_required:
        suggestions = []
        for col in sorted(missing_required):
            matches = get_close_matches(col, list(found), n=1, cutoff=0.6)
            if matches:
                suggestions.append(
                    f'  - "{col}" not found (did you mean "{matches[0]}"?)'
                )
            else:
                suggestions.append(f'  - "{col}" not found')
        raise IndexImportError(
            "Spreadsheet is missing required column(s):\n"
            + "\n".join(suggestions)
            + f"\n\nFound columns: {', '.join(sorted(found))}"
            + f"\nExpected: {', '.join(sorted(KNOWN_COLUMNS))}"
        )

    warnings = []
    missing_optional = (KNOWN_COLUMNS - REQUIRED_COLUMNS) - found
    for col in sorted(missing_optional):
        matches = get_close_matches(col, list(found), n=1, cutoff=0.6)
        hint = f' (did you mean "{matches[0]}"?)' if matches else ""
        warnings.append(f'Optional column "{col}" not found{hint}')
    return warnings


# ---------------------------------------------------------------------------
# Artist merging
# ---------------------------------------------------------------------------


def _artist_merge_key(
    title: Optional[str],
    first_name: Optional[str],
    last_name: Optional[str],
    quals: Optional[str],
) -> str:
    """Build a key for merging duplicate rows that have no courtesy
    distinction. Uses normalised name fields."""
    parts = [
        (title or "").strip().lower(),
        (first_name or "").strip().lower(),
        (last_name or "").strip().lower(),
        (quals or "").strip().lower(),
    ]
    return "|".join(parts)


# ---------------------------------------------------------------------------
# Workbook parsing helper (shared by import & reimport)
# ---------------------------------------------------------------------------


def _open_and_parse_index_workbook(file_path: str) -> Tuple[List[dict], List[str]]:
    """Open an xlsx file and return (raw_rows, header_warnings).

    Each element of *raw_rows* is a dict with normalised fields plus the
    original ``raw_*`` values and a ``cat_nos`` list.  This does NOT touch
    the database — it only parses the file.
    """
    try:
        workbook = load_workbook(filename=file_path, data_only=True)
    except InvalidFileException:
        raise IndexImportError("The uploaded file is not a valid Excel (.xlsx) file.")
    except Exception as exc:
        raise IndexImportError(f"Could not read the uploaded file: {exc}")

    sheet = workbook.active
    if sheet is None or sheet.max_row is None or sheet.max_row < 1:
        raise IndexImportError("The spreadsheet is empty (no rows found).")

    headers = [str(cell.value).strip() if cell.value else "" for cell in sheet[1]]
    header_warnings = _validate_headers(headers)

    def _cell_value(cell):
        v = cell.value
        if cell.quotePrefix and isinstance(v, str):
            v = "'" + v
        return v

    raw_rows: List[dict] = []
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        row_dict = {h: _cell_value(c) for h, c in zip(headers, row)}

        raw_title = row_dict.get("Title")
        raw_first = row_dict.get("First Name")
        raw_last = row_dict.get("Last Name")
        raw_quals = row_dict.get("Quals")
        raw_company = row_dict.get("Company")
        raw_address = row_dict.get("Address 1")
        raw_cat_nos = row_dict.get("Cat Nos")

        if not any(
            [
                raw_title,
                raw_first,
                raw_last,
                raw_quals,
                raw_company,
                raw_address,
                raw_cat_nos,
            ]
        ):
            continue

        title = str(raw_title).strip() if raw_title else None
        first_name = str(raw_first).strip() if raw_first else None
        last_name = str(raw_last).strip() if raw_last else None
        quals = str(raw_quals).strip().rstrip() if raw_quals else None
        company = str(raw_company).strip() if raw_company else None
        address = str(raw_address).strip() if raw_address else None

        if raw_last is not None and not isinstance(raw_last, str):
            last_name = str(raw_last)
            raw_last = str(raw_last)

        if raw_first is not None and not isinstance(raw_first, str):
            first_name = str(raw_first)
            raw_first = str(raw_first)

        cat_nos = parse_cat_nos(str(raw_cat_nos) if raw_cat_nos is not None else None)

        raw_rows.append(
            {
                "row_number": row_idx,
                "raw_title": str(raw_title) if raw_title is not None else None,
                "raw_first_name": str(raw_first) if raw_first is not None else None,
                "raw_last_name": str(raw_last) if raw_last is not None else None,
                "raw_quals": str(raw_quals) if raw_quals is not None else None,
                "raw_company": str(raw_company) if raw_company is not None else None,
                "raw_address": str(raw_address) if raw_address is not None else None,
                "title": title,
                "first_name": first_name,
                "last_name": last_name,
                "quals": quals,
                "company": company,
                "address": address,
                "cat_nos": cat_nos,
            }
        )

    return raw_rows, header_warnings


# ---------------------------------------------------------------------------
# Artist creation from parsed rows (shared by import & reimport)
# ---------------------------------------------------------------------------


def _create_artists_from_parsed_rows(
    db: Session,
    import_record: Import,
    raw_rows: List[dict],
) -> int:
    """Merge and create IndexArtist + IndexCatNumber records from *raw_rows*.

    Returns the number of artist entries created.
    """
    artist_groups: Dict[str, List[dict]] = defaultdict(list)
    for row in raw_rows:
        key = _artist_merge_key(
            row["title"], row["first_name"], row["last_name"], row["quals"]
        )
        artist_groups[key].append(row)

    artist_count = 0
    for key, rows in artist_groups.items():
        courtesy_rows = [r for r in rows if r["address"]]
        no_courtesy_rows = [r for r in rows if not r["address"]]

        if no_courtesy_rows:
            merged = no_courtesy_rows[0].copy()
            merged_cat_entries: List[tuple] = []
            for r in no_courtesy_rows:
                for cn in r["cat_nos"]:
                    merged_cat_entries.append((cn, r["row_number"]))
            _create_artist_entry(
                db, import_record, merged, merged_cat_entries, courtesy=None
            )
            artist_count += 1

            if len(no_courtesy_rows) > 1:
                row_nums = sorted(r["row_number"] for r in no_courtesy_rows)
                rows_str = ", ".join(str(n) for n in row_nums)
                name = f"{no_courtesy_rows[0]['first_name'] or ''} {no_courtesy_rows[0]['last_name'] or ''}".strip()
                db.add(
                    ValidationWarning(
                        import_id=import_record.id,
                        work_id=None,
                        warning_type="duplicate_name_merged",
                        message=f'Rows {rows_str}: Identical name "{name}" merged into one entry',
                    )
                )

        for r in courtesy_rows:
            cat_entries = [(cn, r["row_number"]) for cn in r["cat_nos"]]
            _create_artist_entry(
                db, import_record, r, cat_entries, courtesy=r["address"]
            )
            artist_count += 1

    if artist_count == 0:
        db.add(
            ValidationWarning(
                import_id=import_record.id,
                work_id=None,
                warning_type="empty_spreadsheet",
                message="The spreadsheet has column headers but no data rows.",
            )
        )

    return artist_count


# ---------------------------------------------------------------------------
# Main import function
# ---------------------------------------------------------------------------


def import_index_excel(
    file_path: str,
    db: Session,
    display_name: Optional[str] = None,
) -> Import:
    """Import an Artists' Index spreadsheet.

    Each spreadsheet row becomes an IndexArtist with associated
    IndexCatNumber records.  Rows for the same artist with no courtesy
    distinction are merged (cat numbers combined).
    """
    raw_rows, header_warnings = _open_and_parse_index_workbook(file_path)

    record_name = display_name or file_path

    # Duplicate filename detection
    duplicate_detected = (
        db.query(Import).filter(Import.filename == record_name).first() is not None
    )

    import_record = Import(filename=record_name, product_type="artists_index")
    db.add(import_record)
    db.flush()

    if duplicate_detected:
        db.add(
            ValidationWarning(
                import_id=import_record.id,
                work_id=None,
                warning_type="duplicate_filename",
                message=f"A previous import with filename {record_name!r} already exists",
            )
        )

    for msg in header_warnings:
        db.add(
            ValidationWarning(
                import_id=import_record.id,
                work_id=None,
                warning_type="missing_column",
                message=msg,
            )
        )

    _create_artists_from_parsed_rows(db, import_record, raw_rows)

    db.commit()
    return import_record


def _create_artist_entry(
    db: Session,
    import_record: Import,
    row: dict,
    cat_entries: List[tuple],
    courtesy: Optional[str],
) -> IndexArtist:
    """Create an IndexArtist and its associated IndexCatNumber records.

    cat_entries is a list of (cat_no, source_row) tuples.
    """
    first_name = row["first_name"]
    last_name = row["last_name"]
    quals = row["quals"]
    title = row["title"]
    company_name = row["company"]
    second_artist = None

    # Multi-artist parsing: detect "and X" / "& X" in last_name
    multi = parse_multi_artist(first_name, last_name, quals)
    if multi:
        first_name = multi["first_name"]
        last_name = multi["last_name"]
        quals = multi["quals"]
        second_artist = multi["second_artist"]

    ra_member = is_ra_member(quals)

    is_company_flag = detect_company(first_name, last_name, quals)

    # If detected as company, move last_name into company field
    if is_company_flag and not company_name:
        company_name = last_name

    sort = build_sort_key(last_name, first_name)

    artist = IndexArtist(
        import_id=import_record.id,
        row_number=row["row_number"],
        raw_title=row["raw_title"],
        raw_first_name=row["raw_first_name"],
        raw_last_name=row["raw_last_name"],
        raw_quals=row["raw_quals"],
        raw_company=row["raw_company"],
        raw_address=row["raw_address"],
        title=title,
        first_name=first_name,
        last_name=last_name,
        quals=quals,
        company=company_name,
        second_artist=second_artist,
        is_ra_member=ra_member,
        is_company=is_company_flag,
        sort_key=sort,
    )
    db.add(artist)
    db.flush()

    # Create cat number entries
    for num, src_row in cat_entries:
        db.add(
            IndexCatNumber(
                artist_id=artist.id,
                cat_no=num,
                courtesy=courtesy,
                source_row=src_row,
            )
        )

    # Validation warnings
    if not cat_entries:
        db.add(
            ValidationWarning(
                import_id=import_record.id,
                work_id=None,
                warning_type="missing_cat_nos",
                message=f"Row {row['row_number']}: No catalogue numbers for {first_name or ''} {last_name or ''}".strip(),
            )
        )

    if is_company_flag:
        db.add(
            ValidationWarning(
                import_id=import_record.id,
                work_id=None,
                warning_type="possible_company",
                message=f"Row {row['row_number']}: \"{last_name}\" has no first name — treated as company",
            )
        )

    # Multi-name detection
    if detect_multi_name(first_name, last_name):
        db.add(
            ValidationWarning(
                import_id=import_record.id,
                work_id=None,
                warning_type="multi_artist_name",
                message=f"Row {row['row_number']}: Name may contain multiple artists: {first_name or ''} {last_name or ''}".strip(),
            )
        )

    # Quals-in-name detection
    embedded_qual = detect_quals_in_name(first_name, last_name)
    if embedded_qual:
        db.add(
            ValidationWarning(
                import_id=import_record.id,
                work_id=None,
                warning_type="quals_in_name_field",
                message=f"Row {row['row_number']}: Qualification \"{embedded_qual}\" found in name field: {first_name or ''} {last_name or ''}".strip(),
            )
        )

    # Non-ASCII character detection
    _non_ascii_fields = {
        "last_name": last_name,
        "first_name": first_name,
        "quals": quals,
        "title": title,
        "company": company_name,
        "second_artist": second_artist,
    }
    non_ascii_hits = []
    for field_name, value in _non_ascii_fields.items():
        if not value:
            continue
        chars = sorted({ch for ch in value if ord(ch) > 127}, key=ord)
        if chars:
            samples = ", ".join(f"{ch!r} (U+{ord(ch):04X})" for ch in chars[:5])
            non_ascii_hits.append(f"{field_name}: {samples}")
    if non_ascii_hits:
        db.add(
            ValidationWarning(
                import_id=import_record.id,
                work_id=None,
                warning_type="non_ascii_characters",
                message=(
                    f"Row {row['row_number']}: Non-ASCII characters will be "
                    "unicode-escaped in export — " + "; ".join(non_ascii_hits)
                ),
            )
        )

    return artist


# ---------------------------------------------------------------------------
# Re-import (replace spreadsheet, preserving overrides)
# ---------------------------------------------------------------------------

# Override columns that should be snapshotted / restored across reimports.
_INDEX_OVERRIDE_FIELDS = [
    "is_company_override",
    "first_name_override",
    "last_name_override",
    "title_override",
    "quals_override",
    "second_artist_override",
]


def _artist_identity_key(artist: IndexArtist) -> str:
    """Build a composite key for matching artists across reimports.

    Uses ``sort_key`` (accent-stripped, lowercase last+first) combined
    with the first courtesy address (or empty) so that courtesy vs.
    non-courtesy entries for the same name are treated as distinct.
    """
    # Find courtesy from the artist's cat numbers — courtesy rows store one
    # address value.  We use the first non-None courtesy found.
    courtesy = ""
    if hasattr(artist, "_reimport_courtesy"):
        # fast-path: populated during snapshot
        courtesy = artist._reimport_courtesy or ""
    return f"{artist.sort_key or ''}|{courtesy}"


def reimport_index_excel(
    import_id: _uuid.UUID,
    file_path: str,
    db: Session,
    display_name: Optional[str] = None,
) -> Tuple[Import, Dict[str, int]]:
    """Re-import an Artists' Index spreadsheet into an existing Import,
    preserving overrides and ``include_in_export`` flags.

    Artists are matched by ``sort_key`` + courtesy.  Returns
    ``(import_record, stats)`` with keys: ``matched``, ``added``,
    ``removed``, ``overrides_preserved``.
    """
    import_record = db.query(Import).filter(Import.id == import_id).first()
    if import_record is None:
        raise IndexImportError("Import not found.")

    # 1. Parse new file — fail fast before touching existing data
    raw_rows, header_warnings = _open_and_parse_index_workbook(file_path)

    # 2. Snapshot existing overrides + include_in_export, keyed by identity
    existing_artists = (
        db.query(IndexArtist).filter(IndexArtist.import_id == import_id).all()
    )
    artist_ids = [a.id for a in existing_artists]

    # Load overrides
    existing_overrides: Dict[str, IndexArtistOverride] = {}
    if artist_ids:
        existing_overrides = {
            str(o.artist_id): o
            for o in db.query(IndexArtistOverride)
            .filter(IndexArtistOverride.artist_id.in_(artist_ids))
            .all()
        }

    # Load courtesy per artist (first courtesy from cat numbers)
    courtesy_map: Dict[str, str] = {}
    if artist_ids:
        cat_numbers = (
            db.query(IndexCatNumber)
            .filter(IndexCatNumber.artist_id.in_(artist_ids))
            .all()
        )
        for cn in cat_numbers:
            aid = str(cn.artist_id)
            if aid not in courtesy_map and cn.courtesy:
                courtesy_map[aid] = cn.courtesy

    # Build preservation map: identity_key → {include_in_export, override_fields}
    preserve: Dict[str, dict] = {}
    for a in existing_artists:
        a._reimport_courtesy = courtesy_map.get(str(a.id), "")
        key = _artist_identity_key(a)
        entry: dict = {"include_in_export": a.include_in_export}
        ovr = existing_overrides.get(str(a.id))
        if ovr:
            entry["override"] = {f: getattr(ovr, f) for f in _INDEX_OVERRIDE_FIELDS}
        preserve[key] = entry

    # 3. Delete old data (CASCADE will handle cat_numbers, overrides,
    #    artist-level audit logs)
    if artist_ids:
        db.query(IndexArtistOverride).filter(
            IndexArtistOverride.artist_id.in_(artist_ids)
        ).delete(synchronize_session=False)
        db.query(IndexCatNumber).filter(
            IndexCatNumber.artist_id.in_(artist_ids)
        ).delete(synchronize_session=False)
    db.query(ValidationWarning).filter(ValidationWarning.import_id == import_id).delete(
        synchronize_session=False
    )
    db.query(IndexArtist).filter(IndexArtist.import_id == import_id).delete(
        synchronize_session=False
    )
    db.flush()

    # 4. Import-level warnings (header issues)
    for msg in header_warnings:
        db.add(
            ValidationWarning(
                import_id=import_id,
                work_id=None,
                warning_type="missing_column",
                message=msg,
            )
        )

    # 5. Re-create artists from new spreadsheet
    _create_artists_from_parsed_rows(db, import_record, raw_rows)
    db.flush()

    # 6. Restore overrides + include_in_export for matched artists
    new_artists = db.query(IndexArtist).filter(IndexArtist.import_id == import_id).all()
    # Build courtesy map for new artists
    new_cat_numbers = (
        (
            db.query(IndexCatNumber)
            .filter(IndexCatNumber.artist_id.in_([a.id for a in new_artists]))
            .all()
        )
        if new_artists
        else []
    )
    new_courtesy_map: Dict[str, str] = {}
    for cn in new_cat_numbers:
        aid = str(cn.artist_id)
        if aid not in new_courtesy_map and cn.courtesy:
            new_courtesy_map[aid] = cn.courtesy

    stats = {"matched": 0, "added": 0, "removed": 0, "overrides_preserved": 0}
    matched_keys: set = set()

    for a in new_artists:
        a._reimport_courtesy = new_courtesy_map.get(str(a.id), "")
        key = _artist_identity_key(a)
        if key in preserve and key not in matched_keys:
            matched_keys.add(key)
            entry = preserve[key]
            a.include_in_export = entry["include_in_export"]
            if "override" in entry:
                ovr = IndexArtistOverride(artist_id=a.id, **entry["override"])
                db.add(ovr)
                stats["overrides_preserved"] += 1
            stats["matched"] += 1
        else:
            stats["added"] += 1

    stats["removed"] = len(set(preserve.keys()) - matched_keys)

    # 7. Update import record filename if provided
    if display_name:
        import_record.filename = display_name

    # 8. Audit log entry
    db.add(
        AuditLog(
            import_id=import_id,
            artist_id=None,
            action="reimport",
            field=None,
            old_value=None,
            new_value=(
                f"matched={stats['matched']}, added={stats['added']}, "
                f"removed={stats['removed']}, "
                f"overrides_preserved={stats['overrides_preserved']}"
            ),
        )
    )

    db.commit()
    return import_record, stats
