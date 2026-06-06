import uuid as _uuid
from difflib import get_close_matches
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy.orm import Session

from backend.app.models.audit_log_model import AuditLog
from backend.app.models.import_model import Import
from backend.app.models.override_model import WorkOverride
from backend.app.models.section_model import Section
from backend.app.models.validation_warning_model import ValidationWarning
from backend.app.models.work_model import Work
from backend.app.services.normalisation_service import (
    collect_work_warnings,
    normalise_work,
)
from backend.app.services.reimport_matcher import (
    MatchPlan,
    NewWorkRow,
    OldWorkSnapshot,
    compute_fingerprint,
    match_overrides,
)

# ---------------------------------------------------------------------------
# Expected column headers
# ---------------------------------------------------------------------------

REQUIRED_COLUMNS = {"Cat No", "Title", "Artist"}
KNOWN_COLUMNS = {
    "Cat No",
    "Gallery",
    "Title",
    "Artist",
    "Price",
    "Edition",
    "Artwork",
    "Medium",
}


class ImportError(Exception):
    """Raised when the uploaded file cannot be imported."""

    pass


# ---------------------------------------------------------------------------
# Header validation
# ---------------------------------------------------------------------------


def _validate_headers(headers: list[str]) -> list[str]:
    """Validate spreadsheet headers. Returns a list of warning strings for
    missing optional columns.  Raises ImportError for fatal problems."""

    # Strip empties
    found = {h for h in headers if h}

    if not found:
        raise ImportError(
            "The spreadsheet has no column headers in row 1. "
            "Expected columns: " + ", ".join(sorted(KNOWN_COLUMNS))
        )

    # Check for required columns
    missing_required = REQUIRED_COLUMNS - found
    if missing_required:
        # Try to suggest close matches from what was found
        suggestions = []
        for col in sorted(missing_required):
            matches = get_close_matches(col, list(found), n=1, cutoff=0.6)
            if matches:
                suggestions.append(f'  - "{col}" not found (did you mean "{matches[0]}"?)')
            else:
                suggestions.append(f'  - "{col}" not found')

        raise ImportError(
            "Spreadsheet is missing required column(s):\n"
            + "\n".join(suggestions)
            + f"\n\nFound columns: {', '.join(sorted(found))}"
            + f"\nExpected: {', '.join(sorted(KNOWN_COLUMNS))}"
        )

    # Warnings for missing optional columns
    warnings = []
    missing_optional = (KNOWN_COLUMNS - REQUIRED_COLUMNS) - found
    for col in sorted(missing_optional):
        matches = get_close_matches(col, list(found), n=1, cutoff=0.6)
        hint = f' (did you mean "{matches[0]}"?)' if matches else ""
        warnings.append(f'Optional column "{col}" not found{hint}')

    return warnings


def import_excel(
    file_path: str,
    db: Session,
    honorific_tokens: Optional[List[str]] = None,
    display_name: Optional[str] = None,
    edition_suppress_max: int = 0,
    text_substitutions: Optional[List[dict]] = None,
    title_case_exceptions: Optional[List[str]] = None,
) -> Import:
    # --- Open workbook (catch corrupt / non-Excel files) ---
    try:
        workbook = load_workbook(filename=file_path, data_only=True)
    except InvalidFileException:
        raise ImportError("The uploaded file is not a valid Excel (.xlsx) file.")
    except Exception as exc:
        raise ImportError(f"Could not read the uploaded file: {exc}")

    sheet = workbook.active
    if sheet is None or sheet.max_row is None or sheet.max_row < 1:
        raise ImportError("The spreadsheet is empty (no rows found).")

    # --- Read & validate headers ---
    headers = [str(cell.value).strip() if cell.value else "" for cell in sheet[1]]
    header_warnings = _validate_headers(headers)

    # Use the user-facing display name for the record, falling back to file_path
    record_name = display_name or file_path

    # Duplicate filename detection
    duplicate_detected = db.query(Import).filter(Import.filename == record_name).first() is not None

    import_record = Import(filename=record_name)
    db.add(import_record)
    db.flush()

    # Import-level warning for duplicate filename
    if duplicate_detected:
        db.add(
            ValidationWarning(
                import_id=import_record.id,
                work_id=None,
                warning_type="duplicate_filename",
                message=f"A previous import with filename {file_path!r} already exists",
            )
        )

    # Import-level warnings for missing optional columns
    for msg in header_warnings:
        db.add(
            ValidationWarning(
                import_id=import_record.id,
                work_id=None,
                warning_type="missing_column",
                message=msg,
            )
        )

    sections_map = {}
    section_positions = {}

    def _cell_value(cell):
        """Return the cell's value, restoring a leading apostrophe that Excel
        silently strips when quotePrefix is set on the cell style."""
        v = cell.value
        if cell.quotePrefix and isinstance(v, str):
            v = "'" + v
        return v

    for row in sheet.iter_rows(min_row=2):
        row_dict = {h: _cell_value(c) for h, c in zip(headers, row)}

        raw_cat_no = row_dict.get("Cat No")
        raw_gallery = row_dict.get("Gallery")
        raw_title = row_dict.get("Title")
        raw_artist = row_dict.get("Artist")
        raw_price = row_dict.get("Price")
        raw_edition = row_dict.get("Edition")
        raw_artwork = row_dict.get("Artwork")
        raw_medium = row_dict.get("Medium")

        gallery_name = raw_gallery or "Uncategorised"

        if gallery_name not in sections_map:
            section = Section(
                import_id=import_record.id,
                name=gallery_name,
                position=len(sections_map) + 1,
            )
            db.add(section)
            db.flush()
            sections_map[gallery_name] = section
            section_positions[gallery_name] = 0

        section = sections_map[gallery_name]
        section_positions[gallery_name] += 1

        work = Work(
            import_id=import_record.id,
            section_id=section.id,
            position_in_section=section_positions[gallery_name],
            raw_cat_no=raw_cat_no,
            raw_gallery=raw_gallery,
            raw_title=raw_title,
            raw_artist=raw_artist,
            raw_price=raw_price,
            raw_edition=raw_edition,
            raw_artwork=raw_artwork,
            raw_medium=raw_medium,
        )

        db.add(work)
        normalise_work(
            work,
            honorific_tokens=honorific_tokens,
            edition_suppress_max=edition_suppress_max,
            text_substitutions=text_substitutions,
            title_case_exceptions=title_case_exceptions,
        )
        db.flush()  # ensures work.id is assigned before referencing it in warnings

        # Collect and store work-level validation warnings
        for warning_type, message in collect_work_warnings(
            work,
            edition_suppress_max=edition_suppress_max,
            title_case_exceptions=title_case_exceptions,
        ):
            db.add(
                ValidationWarning(
                    import_id=import_record.id,
                    work_id=work.id,
                    warning_type=warning_type,
                    message=message,
                )
            )

    # Warn if spreadsheet had headers but no data rows
    if not sections_map:
        db.add(
            ValidationWarning(
                import_id=import_record.id,
                work_id=None,
                warning_type="empty_spreadsheet",
                message="The spreadsheet has column headers but no data rows.",
            )
        )

    db.commit()
    return import_record


# ---------------------------------------------------------------------------
# Re-import: update existing import preserving overrides
# ---------------------------------------------------------------------------


def _open_and_parse_workbook(file_path: str) -> Tuple[List[str], List[dict], List[str]]:
    """Open an Excel file, validate headers, and return (headers, rows, header_warnings).

    Each row is a dict mapping header name → cell value.
    Raises ImportError on validation failure.
    """
    try:
        workbook = load_workbook(filename=file_path, data_only=True)
    except InvalidFileException:
        raise ImportError("The uploaded file is not a valid Excel (.xlsx) file.")
    except Exception as exc:
        raise ImportError(f"Could not read the uploaded file: {exc}")

    sheet = workbook.active
    if sheet is None or sheet.max_row is None or sheet.max_row < 1:
        raise ImportError("The spreadsheet is empty (no rows found).")

    headers = [str(cell.value).strip() if cell.value else "" for cell in sheet[1]]
    header_warnings = _validate_headers(headers)

    def _cell_value(cell):
        v = cell.value
        if cell.quotePrefix and isinstance(v, str):
            v = "'" + v
        return v

    rows = [{h: _cell_value(c) for h, c in zip(headers, row)} for row in sheet.iter_rows(min_row=2)]

    return headers, rows, header_warnings


OVERRIDE_FIELDS = [
    "title_override",
    "title_cased_override",
    "artist_name_override",
    "artist_honorifics_override",
    "price_numeric_override",
    "price_text_override",
    "edition_total_override",
    "edition_price_numeric_override",
    "artwork_override",
    "medium_override",
    "notes",
]


def reimport_excel(
    import_id: _uuid.UUID,
    file_path: str,
    db: Session,
    honorific_tokens: Optional[List[str]] = None,
    display_name: Optional[str] = None,
    edition_suppress_max: int = 0,
    text_substitutions: Optional[List[dict]] = None,
    title_case_exceptions: Optional[List[str]] = None,
    gallery_scope: Optional[set] = None,
    dry_run: bool = False,
) -> Tuple[Import, MatchPlan]:
    """Re-import a spreadsheet into an existing Import, preserving overrides.

    Override preservation goes through ``reimport_matcher.match_overrides``
    which combines a cat-no match (gated on fingerprint agreement, so a
    renumbered cat-no can't silently transplant an override onto a different
    work) with a fingerprint fallback (so a work whose cat-no shifted but
    whose content is unchanged keeps its override).

    Parameters
    ----------
    gallery_scope
        When provided, the operation is **scoped**: only galleries whose
        name is in this set are deleted-and-rebuilt; works in any other
        gallery stay physically untouched. New-spreadsheet rows for
        out-of-scope galleries are ignored. Cross-gallery moves into/out of
        the scope are detected and surfaced as warnings in the returned plan.
    dry_run
        When True, parse + compute the plan and return it without mutating
        the DB. The transaction is rolled back at the end so any flushed
        snapshot state is discarded.

    Returns
    -------
    ``(import_record, MatchPlan)``. The plan carries the per-gallery
    summary, counts, and per-finding details (unmatched, ambiguous,
    cross-gallery-move warnings).
    """

    import_record = db.query(Import).filter(Import.id == import_id).first()
    if import_record is None:
        raise ImportError("Import not found.")

    # 1. Parse & validate new file — fail fast before touching existing data
    _headers, rows, header_warnings = _open_and_parse_workbook(file_path)

    # 2. Snapshot existing works + overrides as inputs to the matcher
    existing_works = db.query(Work).filter(Work.import_id == import_id).all()
    work_ids = [w.id for w in existing_works]
    existing_overrides_by_work: Dict = {}
    if work_ids:
        existing_overrides_by_work = {
            o.work_id: o
            for o in db.query(WorkOverride).filter(WorkOverride.work_id.in_(work_ids)).all()
        }

    # Side-table for restoring overrides keyed by old cat_no (the matcher
    # tells us which OLD cat_no maps to which NEW row; we re-create works
    # in document order, then look up by the matcher's mapping).
    preserve_data: Dict[str, dict] = {}
    old_snapshots: List[OldWorkSnapshot] = []
    for w in existing_works:
        cat_key = str(w.raw_cat_no).strip() if w.raw_cat_no is not None else ""
        ovr_obj = existing_overrides_by_work.get(w.id)
        ovr_dict = {f: getattr(ovr_obj, f) for f in OVERRIDE_FIELDS} if ovr_obj else None
        old_snapshots.append(
            OldWorkSnapshot(
                cat_no=cat_key,
                gallery=str(w.raw_gallery or ""),
                fingerprint=compute_fingerprint(w.raw_title, w.raw_artist, w.raw_medium),
                include_in_export=w.include_in_export,
                override=ovr_dict,
                raw_title=w.raw_title,
                raw_artist=w.raw_artist,
            )
        )
        # Last-write-wins by cat_no for restore (mirrors the historical
        # behaviour for the pathological "two old works with the same
        # cat_no" case — should be vanishingly rare in real data).
        if cat_key:
            preserve_data[cat_key] = {
                "include_in_export": w.include_in_export,
                "override": ovr_dict,
            }

    # 3. Build new-side rows (no DB writes yet — the matcher needs to see
    # the spreadsheet first so we can decide before mutating).
    new_rows: List[NewWorkRow] = []
    for row_dict in rows:
        raw_cat_no = row_dict.get("Cat No")
        gallery_name = row_dict.get("Gallery") or "Uncategorised"
        new_rows.append(
            NewWorkRow(
                cat_no=(str(raw_cat_no).strip() if raw_cat_no is not None else ""),
                gallery=str(gallery_name),
                fingerprint=compute_fingerprint(
                    row_dict.get("Title"),
                    row_dict.get("Artist"),
                    row_dict.get("Medium"),
                ),
                raw_title=row_dict.get("Title"),
                raw_artist=row_dict.get("Artist"),
            )
        )

    # 4. Compute the plan (pure, no side effects)
    plan = match_overrides(old_snapshots, new_rows, gallery_scope=gallery_scope)

    # 5. Dry-run: return the plan without touching the DB
    if dry_run:
        # Roll back any implicit reads so the next call sees a clean session.
        db.rollback()
        return import_record, plan

    # 6. Delete the scope. Two paths:
    #    - Full re-import (scope None): wipe everything for this import.
    #    - Selective: only delete sections in the scope; out-of-scope
    #      sections, works, overrides, and per-work warnings stay intact.
    section_positions: Dict[str, int] = {}
    sections_map: Dict[str, Section] = {}
    preserved_section_positions: Dict[str, int] = {}

    if gallery_scope is None:
        if work_ids:
            db.query(WorkOverride).filter(WorkOverride.work_id.in_(work_ids)).delete(
                synchronize_session=False
            )
        db.query(ValidationWarning).filter(ValidationWarning.import_id == import_id).delete(
            synchronize_session=False
        )
        db.query(Work).filter(Work.import_id == import_id).delete(synchronize_session=False)
        db.query(Section).filter(Section.import_id == import_id).delete(synchronize_session=False)
        db.flush()
        next_position = 1
    else:
        in_scope_sections = (
            db.query(Section)
            .filter(
                Section.import_id == import_id,
                Section.name.in_(gallery_scope),
            )
            .all()
        )
        # Remember each in-scope section's position so we can put the
        # re-created section back in the same slot — UI ordering stays stable.
        for s in in_scope_sections:
            preserved_section_positions[s.name] = s.position
        in_scope_section_ids = [s.id for s in in_scope_sections]
        in_scope_work_ids = [
            w.id for w in existing_works if w.section_id in set(in_scope_section_ids)
        ]
        if in_scope_work_ids:
            db.query(WorkOverride).filter(WorkOverride.work_id.in_(in_scope_work_ids)).delete(
                synchronize_session=False
            )
            # Per-work validation warnings only (header warnings keyed by
            # work_id=NULL are global and survive a selective re-import).
            db.query(ValidationWarning).filter(
                ValidationWarning.work_id.in_(in_scope_work_ids)
            ).delete(synchronize_session=False)
        if in_scope_section_ids:
            db.query(Work).filter(Work.section_id.in_(in_scope_section_ids)).delete(
                synchronize_session=False
            )
            db.query(Section).filter(Section.id.in_(in_scope_section_ids)).delete(
                synchronize_session=False
            )
        db.flush()
        # New galleries (in scope but no pre-existing section) get
        # positions after the current max.
        current_max = (
            db.query(Section.position)
            .filter(Section.import_id == import_id)
            .order_by(Section.position.desc())
            .limit(1)
            .scalar()
            or 0
        )
        next_position = current_max + 1

    # 7. Re-create sections + works (only those in scope).
    # Header warnings only on full re-import — they apply to the spreadsheet
    # as a whole and would duplicate noise on every selective re-upload.
    if gallery_scope is None:
        for msg in header_warnings:
            db.add(
                ValidationWarning(
                    import_id=import_id,
                    work_id=None,
                    warning_type="missing_column",
                    message=msg,
                )
            )

    # Map (in_scope_row_index → matched item) so we can restore override
    # data as each new Work is created.
    matched_by_in_scope_idx: Dict[int, object] = {m.new_row_index: m for m in plan.matched}
    in_scope_counter = 0

    for row_dict in rows:
        raw_cat_no = row_dict.get("Cat No")
        raw_gallery = row_dict.get("Gallery")
        raw_title = row_dict.get("Title")
        raw_artist = row_dict.get("Artist")
        raw_price = row_dict.get("Price")
        raw_edition = row_dict.get("Edition")
        raw_artwork = row_dict.get("Artwork")
        raw_medium = row_dict.get("Medium")

        gallery_name = raw_gallery or "Uncategorised"

        # Skip rows outside the requested scope
        if gallery_scope is not None and gallery_name not in gallery_scope:
            continue

        if gallery_name not in sections_map:
            # Reuse the original position when possible so the UI keeps
            # showing galleries in the original order; otherwise append.
            position = preserved_section_positions.get(gallery_name)
            if position is None:
                position = next_position
                next_position += 1
            section = Section(
                import_id=import_id,
                name=gallery_name,
                position=position,
            )
            db.add(section)
            db.flush()
            sections_map[gallery_name] = section
            section_positions[gallery_name] = 0

        section = sections_map[gallery_name]
        section_positions[gallery_name] += 1

        work = Work(
            import_id=import_id,
            section_id=section.id,
            position_in_section=section_positions[gallery_name],
            raw_cat_no=raw_cat_no,
            raw_gallery=raw_gallery,
            raw_title=raw_title,
            raw_artist=raw_artist,
            raw_price=raw_price,
            raw_edition=raw_edition,
            raw_artwork=raw_artwork,
            raw_medium=raw_medium,
        )
        db.add(work)
        normalise_work(
            work,
            honorific_tokens=honorific_tokens,
            edition_suppress_max=edition_suppress_max,
            text_substitutions=text_substitutions,
            title_case_exceptions=title_case_exceptions,
        )
        db.flush()

        # Restore override + include_in_export when the matcher paired this
        # row with an old work.
        matched_item = matched_by_in_scope_idx.get(in_scope_counter)
        if matched_item is not None:
            preserved = preserve_data.get(matched_item.old_cat_no)
            if preserved is not None:
                work.include_in_export = preserved["include_in_export"]
                if preserved["override"]:
                    db.add(WorkOverride(work_id=work.id, **preserved["override"]))
        in_scope_counter += 1

        # Work-level validation warnings
        for warning_type, message in collect_work_warnings(
            work,
            edition_suppress_max=edition_suppress_max,
            title_case_exceptions=title_case_exceptions,
        ):
            db.add(
                ValidationWarning(
                    import_id=import_id,
                    work_id=work.id,
                    warning_type=warning_type,
                    message=message,
                )
            )

    if not sections_map and gallery_scope is None:
        db.add(
            ValidationWarning(
                import_id=import_id,
                work_id=None,
                warning_type="empty_spreadsheet",
                message="The spreadsheet has column headers but no data rows.",
            )
        )

    # 8. Update display name only on a full re-import — selective uploads
    # don't represent a new master file.
    if display_name and gallery_scope is None:
        import_record.filename = display_name

    # 9. Audit log entry with the full plan summary
    audit_parts = [
        f"matched={plan.matched_by_cat_no + plan.matched_by_fingerprint}",
        f"added={plan.added}",
        f"removed={plan.removed}",
        f"overrides_preserved={plan.overrides_preserved}",
        f"matched_by_cat_no={plan.matched_by_cat_no}",
        f"matched_by_fingerprint={plan.matched_by_fingerprint}",
        f"overrides_at_risk={plan.overrides_at_risk}",
    ]
    if gallery_scope:
        audit_parts.append(f"scope={sorted(gallery_scope)}")
    db.add(
        AuditLog(
            import_id=import_id,
            work_id=None,
            action="reimport",
            field=None,
            old_value=None,
            new_value=", ".join(audit_parts),
        )
    )

    db.commit()
    return import_record, plan
